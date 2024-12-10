import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import pandas as pd
from concurrent.futures import ProcessPoolExecutor
from selenium.webdriver.common.by import By
from openpyxl.styles import Alignment, Font, Border, Side


def setup_driver():
    """设置并返回WebDriver实例"""
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    service = Service(executable_path="C:\\soft\\chromedriver-win64\\chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver


def fetch_blog_data(url):
    """抓取网页内容"""
    driver = setup_driver()
    driver.get(url)
    time.sleep(3)  # 等待页面加载

    # 获取页面标题作为文件名
    try:
        page_title = driver.find_element(By.TAG_NAME, 'h1').text.strip()
    except Exception as e:
        page_title = "Unknown"

    links = []  # 采用列表保证顺序
    titles = []  # 采用列表保证顺序

    # 获取所有的 <a> 标签
    elements = driver.find_elements(By.CSS_SELECTOR, "a")
    for element in elements:
        title = element.get_attribute("title") or element.text.strip()  # 优先使用 title，如果没有则使用文本
        link = element.get_attribute("href")

        # 过滤掉无效链接和无标题链接
        if link and not link.startswith(("javascript:", "void(0)")) and title:
            links.append(link)
            titles.append(title)

    driver.quit()

    return page_title, list(zip(titles, links))  # 返回标题和链接的配对


def save_to_excel(data, filename):
    """将链接和标题保存到 Excel 文件"""
    df = pd.DataFrame(data, columns=["标题", "链接"])

    # 移除无效数据
    df = df[~df['链接'].str.startswith(("javascript:", "void(0)"))]  # 过滤无效链接

    # 插入序号列
    df.insert(0, '序号', range(1, len(df) + 1))

    # 保存到 Excel
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Links")

        # 获取 DataFrame 的工作表
        workbook = writer.book
        worksheet = workbook["Links"]

        # 设置标题行加粗
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')

        # 设置所有单元格内容的对齐方式（水平方向左对齐，垂直方向居中）
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # 设置列宽自适应
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列名
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # 适当增加宽度
            worksheet.column_dimensions[column].width = adjusted_width

        # 设置边框
        border = Border(left=Side(border_style="thin"),
                        right=Side(border_style="thin"),
                        top=Side(border_style="thin"),
                        bottom=Side(border_style="thin"))

        # 应用边框到数据区域
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border

        # 应用边框到标题行
        for cell in worksheet[1]:
            cell.border = border

    print(f"数据已保存到 {filename}")


def process_url(url):
    """处理单个URL，保存数据到Excel"""
    print(f"开始抓取: {url}")
    page_title, title_link_pairs = fetch_blog_data(url)

    # 根据页面标题创建文件名
    filename = f"{page_title}.xlsx"

    # 保存到 Excel
    save_to_excel(title_link_pairs, filename)


def main():
    """主程序入口"""
    with open("urls.txt", "r") as file:
        urls = [line.strip() for line in file.readlines() if line.strip()]

    # 使用多进程并发处理 URL
    with ProcessPoolExecutor() as executor:
        executor.map(process_url, urls)


if __name__ == "__main__":
    main()
